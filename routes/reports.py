import sqlite3
import os
from datetime import datetime, timedelta
from flask import (
    Blueprint, render_template, request, redirect,
    url_for, flash, session, current_app
)
from db import get_db
from utils.format import obtener_inicial_dia

 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

 
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from utils.roles import requiere_rol


reports_bp = Blueprint('reports', __name__, url_prefix='/reporte_mensual')

@reports_bp.before_request
def require_login():
    if not session.get('usuario'):
        return redirect(url_for('auth.login'))

@reports_bp.route('/', methods=['GET', 'POST'])
@requiere_rol(1, 2)
def reporte_mensual():
    mes = None
    fechas_semana = []
    fechas_del_mes = []
    datos_docentes = {}
    total_asistencias_global = 0
    total_tardanzas_global = 0
    total_inasistencias_global = 0
    total_justificaciones_global = 0

    if request.method == 'POST':
        mes = request.form.get('mes')
        if mes:
            try:
                desde_dt = datetime.strptime(mes + "-01", "%Y-%m-%d")
                next_month = desde_dt.replace(day=28) + timedelta(days=4)
                hasta_dt = next_month - timedelta(days=next_month.day)

                db = get_db()
                c = db.cursor()

                feriados = set(row['fecha'] for row in c.execute("SELECT fecha FROM fechas_no_laborables WHERE fecha >= ? AND fecha <= ?", (desde_dt.strftime('%Y-%m-%d'), hasta_dt.strftime('%Y-%m-%d'))).fetchall())

                fechas_del_mes = [
                    desde_dt + timedelta(days=i)
                    for i in range((hasta_dt - desde_dt).days + 1)
                    if (desde_dt + timedelta(days=i)).weekday() < 5
                ]

                semana_actual_para_template = []
                for f_template in fechas_del_mes:
                    semana_actual_para_template.append(f_template)
                    if len(semana_actual_para_template) == 5:
                        fechas_semana.append(semana_actual_para_template.copy())
                        semana_actual_para_template.clear()
                if semana_actual_para_template:
                    fechas_semana.append(semana_actual_para_template.copy())

                docs_query = """
                    SELECT d.dni,
                           d.nombres || ' ' || d.apellido_paterno || ' ' || d.apellido_materno AS nombre,
                           COALESCE(cr.nombre_cargo, 'N/A') as nombre_cargo
                    FROM docentes d
                    LEFT JOIN cargos cr ON d.cargo_id = cr.id_cargo
                    ORDER BY cr.nombre_cargo, d.apellido_paterno, d.apellido_materno, d.nombres
                """
                docs = c.execute(docs_query).fetchall()

                for d_docente in docs:
                    dni = d_docente['dni']
                    
                     
                    datos_docentes[dni] = {
                        'nombre': d_docente['nombre'],
                        'cargo': d_docente['nombre_cargo'],
                        'asistencias': {},
                        'total': 0,
                        'tardanzas': 0,
                        'inasistencias': 0,
                        'justificaciones': 0,  
                        'permisos': 0,         
                        'porcentaje': 0
                    }

                    for f_dia_obj in fechas_del_mes:
                        fecha_str = f_dia_obj.strftime('%Y-%m-%d')
                        dia_nombre_eng = f_dia_obj.strftime('%A')
                        dias_es_map = {
                            "Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miercoles",
                            "Thursday": "Jueves", "Friday": "Viernes"
                        }
                        dia_nombre_bd = dias_es_map.get(dia_nombre_eng, dia_nombre_eng)

                        if fecha_str in feriados:
                            datos_docentes[dni]['asistencias'][f_dia_obj.date()] = "🔕"
                            continue

                        permiso_activo_propio = c.execute("""
                            SELECT 1 FROM permisos_docentes
                            WHERE dni = ? AND ? BETWEEN fecha_inicio AND fecha_fin
                        """, (dni, fecha_str)).fetchone()

                        if permiso_activo_propio:
                            datos_docentes[dni]['asistencias'][f_dia_obj.date()] = "🅿"
                             
                            datos_docentes[dni]['permisos'] += 1 
                            total_justificaciones_global += 1  
                            continue

                        deberia_trabajar_hoy = False
                        horario_propio = c.execute("SELECT 1 FROM horarios WHERE dni = ? AND dia_semana = ?", (dni, dia_nombre_bd)).fetchone()
                        if horario_propio:
                            deberia_trabajar_hoy = True
                        else:
                            sustitucion_cubre = c.execute("""
                                SELECT 1 FROM sustituciones s
                                JOIN permisos_docentes pd_titular ON s.permiso_id = pd_titular.id
                                JOIN horarios h_titular ON pd_titular.dni = h_titular.dni
                                WHERE s.sustituto_dni = ?
                                  AND ? BETWEEN s.fecha_inicio AND s.fecha_fin
                                  AND ? BETWEEN pd_titular.fecha_inicio AND pd_titular.fecha_fin
                                  AND h_titular.dia_semana = ?
                            """, (dni, fecha_str, fecha_str, dia_nombre_bd)).fetchone()
                            if sustitucion_cubre:
                                deberia_trabajar_hoy = True
                        
                        if not deberia_trabajar_hoy:
                            datos_docentes[dni]['asistencias'][f_dia_obj.date()] = "—"
                            continue

                        if f_dia_obj.date() > datetime.now().date():
                            datos_docentes[dni]['asistencias'][f_dia_obj.date()] = "—"
                            continue

                        hay_registro_general = c.execute("SELECT 1 FROM asistencias WHERE DATE(fecha)=? LIMIT 1", (fecha_str,)).fetchone()
                        reg = c.execute("SELECT id, estado FROM asistencias WHERE dni = ? AND DATE(fecha) = ?", (dni, fecha_str)).fetchone()
                        
                        simbolo_dia = "?"
                        if reg:
                            estado_asistencia = reg['estado']
                            justificada_estado_str = None
                            if estado_asistencia in ("Inasistencia", "Tarde"):
                                justif_row = c.execute("SELECT estado FROM justificaciones WHERE asistencia_id = ?", (reg['id'],)).fetchone()
                                if justif_row: justificada_estado_str = justif_row['estado']
                            
                            if justificada_estado_str == "Aprobada":
                                 
                                simbolo_dia = "J✔"; datos_docentes[dni]['justificaciones'] += 1; total_justificaciones_global += 1
                            elif justificada_estado_str == "Pendiente":
                                simbolo_dia = "J⧗"
                            else:
                                if estado_asistencia == 'A tiempo':
                                    simbolo_dia = "✔"; datos_docentes[dni]['total'] += 1; total_asistencias_global += 1
                                elif estado_asistencia == 'Tarde':
                                    simbolo_dia = "🕒"; datos_docentes[dni]['tardanzas'] += 1; total_tardanzas_global += 1
                                elif estado_asistencia == 'Inasistencia':
                                    simbolo_dia = "✗"; datos_docentes[dni]['inasistencias'] += 1; total_inasistencias_global += 1
                                else: simbolo_dia = "?"
                        elif hay_registro_general:
                            simbolo_dia = "✗"; datos_docentes[dni]['inasistencias'] += 1; total_inasistencias_global += 1
                        else: simbolo_dia = "—"
                        datos_docentes[dni]['asistencias'][f_dia_obj.date()] = simbolo_dia

                     
                    
                     
                    dias_logrados = datos_docentes[dni]['total'] + datos_docentes[dni]['justificaciones']
                    
                     
                    dias_programados_totales = 0
                    for f_eval in fechas_del_mes:
                        simbolo = datos_docentes[dni]['asistencias'].get(f_eval.date())
                        if simbolo not in ["🔕", "—"] and f_eval.date() <= datetime.now().date():
                            dias_programados_totales += 1
                    
                     
                    dias_base_para_calculo = dias_programados_totales - datos_docentes[dni]['permisos']

                    if dias_base_para_calculo > 0:
                        datos_docentes[dni]['porcentaje'] = round((dias_logrados / dias_base_para_calculo) * 100)
                    else:
                         
                        datos_docentes[dni]['porcentaje'] = 100

            except Exception as e:
                current_app.logger.error(f"Error al generar reporte mensual: {e.__class__.__name__}: {e}")
                flash(f"❌ Error al generar reporte: {e.__class__.__name__}: {e}", 'danger')
                mes = None; fechas_semana = []; fechas_del_mes = []; datos_docentes = {}
                total_asistencias_global = 0; total_tardanzas_global = 0
                total_inasistencias_global = 0; total_justificaciones_global = 0
    
    return render_template('reporte_mensual.html',
        fechas_semana=fechas_semana,
        fechas_del_mes=fechas_del_mes,
        docentes=datos_docentes,
        mes=mes,
        total_asistencias_global=total_asistencias_global,
        total_tardanzas_global=total_tardanzas_global,
        total_inasistencias_global=total_inasistencias_global,
        total_justificaciones_global=total_justificaciones_global,
        obtener_inicial_dia=obtener_inicial_dia
    )


@reports_bp.route('/exportar_excel', methods=['POST'])
@requiere_rol(1, 2)
def exportar_excel():
    try:
        mes = request.form.get('mes')
        if not mes:
            flash("⚠️ Debes seleccionar un mes antes de exportar el reporte.", "warning")
            return redirect(url_for('reports.reporte_mensual'))

        desde_dt = datetime.strptime(mes + "-01", "%Y-%m-%d")
        next_month = desde_dt.replace(day=28) + timedelta(days=4)
        hasta_dt = next_month - timedelta(days=next_month.day)

        fechas_del_mes = [
            desde_dt + timedelta(days=i)
            for i in range((hasta_dt - desde_dt).days + 1)
            if (desde_dt + timedelta(days=i)).weekday() < 5
        ]

        db = get_db()
        c = db.cursor()
        
        docs_query = """
            SELECT d.dni, 
                   d.nombres || ' ' || d.apellido_paterno || ' ' || d.apellido_materno AS nombre,
                   COALESCE(cr.nombre_cargo, 'N/A') as nombre_cargo
            FROM docentes d
            LEFT JOIN cargos cr ON d.cargo_id = cr.id_cargo
            ORDER BY nombre_cargo, d.apellido_paterno, d.apellido_materno, d.nombres
        """
        docs = c.execute(docs_query).fetchall()

        feriados = set(
            row['fecha'] for row in c.execute("SELECT fecha FROM fechas_no_laborables WHERE fecha >= ? AND fecha <= ?", (desde_dt.strftime('%Y-%m-%d'), hasta_dt.strftime('%Y-%m-%d'))).fetchall()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = f"Asistencias {mes}"

        encabezado = ["DNI", "Docente", "Cargo"] + [f.strftime('%d-%b') for f in fechas_del_mes] + ["A Tiempo", "Tardanzas", "Justif./Perm.", "Inasistencias", "% Asist."]
        ws.append(encabezado)

        for cell in ws[1]: 
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="BDD7EE")

        g_total_atiempo = 0
        g_total_tardanzas = 0
        g_total_justificaciones_y_permisos = 0  
        g_total_inasistencias = 0

        for d_docente in docs:
            dni = d_docente['dni']
            nombre = d_docente['nombre']
            cargo = d_docente['nombre_cargo']

            fila_excel = [dni, nombre, cargo]
            
             
            d_atiempo = 0
            d_tardanzas = 0
            d_inasistencias = 0
            d_justificaciones = 0  
            d_permisos = 0       
             
            d_dias_laborables_programados = 0 

            for f_dia_obj in fechas_del_mes:
                fecha_str = f_dia_obj.strftime('%Y-%m-%d')
                dia_nombre_eng = f_dia_obj.strftime('%A')
                dias_es_map = {
                    "Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miercoles",
                    "Thursday": "Jueves", "Friday": "Viernes"
                }
                dia_nombre_bd = dias_es_map.get(dia_nombre_eng, dia_nombre_eng)
                
                simbolo_dia_excel = "?"

                if fecha_str in feriados:
                    simbolo_dia_excel = "F"
                    fila_excel.append(simbolo_dia_excel)
                    continue

                permiso_activo_propio = c.execute("""
                    SELECT 1 FROM permisos_docentes
                    WHERE dni = ? AND ? BETWEEN fecha_inicio AND fecha_fin
                """, (dni, fecha_str)).fetchone()

                if permiso_activo_propio:
                    simbolo_dia_excel = "P"
                     
                    d_permisos += 1 
                    fila_excel.append(simbolo_dia_excel)
                     
                    continue

                deberia_trabajar_hoy = False
                horario_propio = c.execute("SELECT 1 FROM horarios WHERE dni=? AND dia_semana=?", (dni, dia_nombre_bd)).fetchone()
                if horario_propio:
                    deberia_trabajar_hoy = True
                else:
                    sustitucion_cubre = c.execute("""
                        SELECT 1 FROM sustituciones s
                        JOIN permisos_docentes pd_titular ON s.permiso_id = pd_titular.id
                        JOIN horarios h_titular ON pd_titular.dni = h_titular.dni
                        WHERE s.sustituto_dni = ?
                          AND ? BETWEEN s.fecha_inicio AND s.fecha_fin
                          AND ? BETWEEN pd_titular.fecha_inicio AND pd_titular.fecha_fin
                          AND h_titular.dia_semana = ?
                    """, (dni, fecha_str, fecha_str, dia_nombre_bd)).fetchone()
                    if sustitucion_cubre:
                        deberia_trabajar_hoy = True
                
                if not deberia_trabajar_hoy:
                    simbolo_dia_excel = "—"
                    fila_excel.append(simbolo_dia_excel)
                    continue
                
                 
                 
                d_dias_laborables_programados += 1 

                hay_registro_general = c.execute("SELECT 1 FROM asistencias WHERE DATE(fecha)=? LIMIT 1", (fecha_str,)).fetchone()
                reg = c.execute("SELECT id, estado FROM asistencias WHERE dni = ? AND DATE(fecha) = ?", (dni, fecha_str)).fetchone()

                if reg:
                    estado_asistencia = reg['estado']
                    justificada_estado_str = None
                    if estado_asistencia in ("Inasistencia", "Tarde"):
                        justif_row = c.execute("SELECT estado FROM justificaciones WHERE asistencia_id = ?", (reg['id'],)).fetchone()
                        if justif_row: justificada_estado_str = justif_row['estado']
                    
                    if justificada_estado_str == "Aprobada":
                        simbolo_dia_excel = "J"
                         
                        d_justificaciones += 1
                    elif justificada_estado_str == "Pendiente":
                        simbolo_dia_excel = "J?"
                    else:
                        if estado_asistencia == 'A tiempo':
                            simbolo_dia_excel = "A"
                            d_atiempo += 1
                        elif estado_asistencia == 'Tarde':
                            simbolo_dia_excel = "T"
                            d_tardanzas += 1
                        elif estado_asistencia == 'Inasistencia':
                            simbolo_dia_excel = "I"
                            d_inasistencias += 1
                elif hay_registro_general:
                    simbolo_dia_excel = "I"
                    d_inasistencias += 1
                else:
                    simbolo_dia_excel = "—"
                     
                    if d_dias_laborables_programados > 0:
                         d_dias_laborables_programados -=1
                
                fila_excel.append(simbolo_dia_excel)
            
             

             
            total_justif_perm_columna = d_justificaciones + d_permisos

             
            numerador_porcentaje = d_atiempo + d_justificaciones
             
            denominador_porcentaje = d_dias_laborables_programados

            if denominador_porcentaje > 0:
                porcentaje_asistencia = round((numerador_porcentaje / denominador_porcentaje) * 100)
            else:
                 
                porcentaje_asistencia = 100
            
            fila_excel += [d_atiempo, d_tardanzas, total_justif_perm_columna, d_inasistencias, f"{porcentaje_asistencia}%"]
            ws.append(fila_excel)

            g_total_atiempo += d_atiempo
            g_total_tardanzas += d_tardanzas
            g_total_justificaciones_y_permisos += total_justif_perm_columna
            g_total_inasistencias += d_inasistencias

        fila_total_global = ["", "Total general", ""]
        fila_total_global += [""] * len(fechas_del_mes)
        fila_total_global += [g_total_atiempo, g_total_tardanzas, g_total_justificaciones_y_permisos, g_total_inasistencias, ""]
        ws.append(fila_total_global)
        
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")

        for col in ws.columns:
            max_len = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_len + 2) if max_len < 30 else 30
            ws.column_dimensions[column].width = adjusted_width

        reportes_dir = current_app.config['REPORTS_PATH']
        os.makedirs(reportes_dir, exist_ok=True)
        output_path = os.path.join(reportes_dir, f"reporte_asistencias_{mes}.xlsx")
        wb.save(output_path)

        flash(f"✅ Reporte Excel generado en: {output_path}", "success")
    except Exception as e:
        current_app.logger.error(f"Error al generar Excel: {e.__class__.__name__}: {e}")
        flash(f"❌ Ocurrió un error al generar el archivo Excel: {e.__class__.__name__}: {e}", "danger")

    return redirect(url_for('reports.reporte_mensual'))


@reports_bp.route('/exportar_pdf', methods=['POST'])
@requiere_rol(1, 2)
def exportar_pdf():
    try:
        mes = request.form.get('mes')
        if not mes:
            flash("⚠️ Debes seleccionar un mes antes de exportar.", "warning")
            return redirect(url_for('reports.reporte_mensual'))

         
        desde_dt = datetime.strptime(mes + "-01", "%Y-%m-%d")
        next_month = desde_dt.replace(day=28) + timedelta(days=4)
        hasta_dt = next_month - timedelta(days=next_month.day)

        fechas_del_mes = [
            desde_dt + timedelta(days=i)
            for i in range((hasta_dt - desde_dt).days + 1)
            if (desde_dt + timedelta(days=i)).weekday() < 5
        ]
        
        db = get_db()

        director_cargo_id_row = db.execute("SELECT id_cargo FROM cargos WHERE nombre_cargo = 'Director'").fetchone()
        director_nombre_completo = "Director no asignado"
        if director_cargo_id_row:
            director_docente_row = db.execute(
                "SELECT nombres || ' ' || apellido_paterno || ' ' || apellido_materno AS nombre FROM docentes WHERE cargo_id = ?", 
                (director_cargo_id_row['id_cargo'],)
            ).fetchone()
            if director_docente_row:
                director_nombre_completo = director_docente_row['nombre'].upper()

        docs_query = """
            SELECT d.dni, d.nombres, d.apellido_paterno, d.apellido_materno,
                   COALESCE(cr.nombre_cargo, 'N/A') as nombre_cargo
            FROM docentes d
            LEFT JOIN cargos cr ON d.cargo_id = cr.id_cargo
            ORDER BY nombre_cargo, d.apellido_paterno, d.apellido_materno, d.nombres
        """
        docentes = db.execute(docs_query).fetchall()

        feriados = set(
            row['fecha'] for row in db.execute("SELECT fecha FROM fechas_no_laborables WHERE fecha >= ? AND fecha <= ?", (desde_dt.strftime('%Y-%m-%d'), hasta_dt.strftime('%Y-%m-%d'))).fetchall()
        )

         
        encabezado_principal_pdf = ["N°", "APELLIDOS Y NOMBRES", "CARGO"]
        encabezado_principal_pdf += [f.strftime('%d') for f in fechas_del_mes]
        encabezado_principal_pdf += ["A Tiempo", "Tard.", "Justif.", "Inasist.", "%"]

        data_para_tabla = [encabezado_principal_pdf]
        
        g_pdf_atiempo, g_pdf_tardanzas, g_pdf_inasistencias, g_pdf_justif_y_perm = 0, 0, 0, 0
        
        styles = getSampleStyleSheet()
        
        for idx, d_docente in enumerate(docentes):
            nombre_completo = f"{d_docente['apellido_paterno']} {d_docente['apellido_materno']}, {d_docente['nombres']}"
            
            fila_data = [str(idx + 1), Paragraph(nombre_completo, styles['Normal']), d_docente['nombre_cargo']]
            
             
            d_atiempo = 0
            d_tardanzas = 0
            d_inasistencias = 0
            d_justificaciones = 0  
            d_permisos = 0       
             
            d_dias_laborables_programados = 0 
            
            for f_dia_obj in fechas_del_mes:
                fecha_str = f_dia_obj.strftime('%Y-%m-%d')
                dia_nombre_bd = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"][f_dia_obj.weekday()]
                simbolo_actual = "?"

                if fecha_str in feriados:
                    simbolo_actual = "F"
                    fila_data.append(simbolo_actual)
                    continue

                permiso_activo = db.execute("""
                    SELECT 1 FROM permisos_docentes
                    WHERE dni = ? AND ? BETWEEN fecha_inicio AND fecha_fin
                """, (d_docente['dni'], fecha_str)).fetchone()
                if permiso_activo:
                    simbolo_actual = "P"
                     
                    d_permisos += 1
                    fila_data.append(simbolo_actual)
                    continue

                deberia_trabajar_hoy = False
                horario_propio = db.execute("SELECT 1 FROM horarios WHERE dni=? AND dia_semana=?", (d_docente['dni'], dia_nombre_bd)).fetchone()
                if horario_propio:
                    deberia_trabajar_hoy = True
                else:
                    sustitucion_cubre = db.execute("""
                        SELECT 1 FROM sustituciones s
                        JOIN permisos_docentes p ON s.permiso_id = p.id
                        JOIN horarios h ON p.dni = h.dni
                        WHERE s.sustituto_dni = ?
                          AND ? BETWEEN s.fecha_inicio AND s.fecha_fin
                          AND ? BETWEEN p.fecha_inicio AND p.fecha_fin
                          AND h.dia_semana = ?
                    """, (d_docente['dni'], fecha_str, fecha_str, dia_nombre_bd)).fetchone()
                    if sustitucion_cubre:
                        deberia_trabajar_hoy = True
                
                if not deberia_trabajar_hoy:
                    simbolo_actual = ""
                    fila_data.append(simbolo_actual)
                    continue
                
                 
                d_dias_laborables_programados += 1
                
                reg = db.execute("SELECT id, estado FROM asistencias WHERE dni=? AND DATE(fecha)=?", (d_docente['dni'], fecha_str)).fetchone()
                if reg:
                    estado_asistencia, just_estado = reg['estado'], None
                    if estado_asistencia in ("Inasistencia", "Tarde"):
                        just_row = db.execute("SELECT estado FROM justificaciones WHERE asistencia_id=?", (reg['id'],)).fetchone()
                        if just_row: just_estado = just_row['estado']
                    
                    if just_estado == "Aprobada":
                        simbolo_actual = "J"
                         
                        d_justificaciones += 1
                    elif just_estado == "Pendiente":
                        simbolo_actual = "J?"
                    elif estado_asistencia == 'A tiempo':
                        simbolo_actual = ""
                        d_atiempo += 1
                    elif estado_asistencia == 'Tarde':
                        simbolo_actual = "T"
                        d_tardanzas += 1
                    elif estado_asistencia == 'Inasistencia':
                        simbolo_actual = "I"
                        d_inasistencias += 1
                elif db.execute("SELECT 1 FROM asistencias WHERE DATE(fecha)=? LIMIT 1", (fecha_str,)).fetchone():
                    simbolo_actual = "I"
                    d_inasistencias += 1
                else:
                    simbolo_actual = ""
                    if d_dias_laborables_programados > 0:
                        d_dias_laborables_programados -= 1
                
                fila_data.append(simbolo_actual)

             

             
            total_justif_perm_columna = d_justificaciones + d_permisos
            
            numerador_porcentaje = d_atiempo + d_justificaciones
            denominador_porcentaje = d_dias_laborables_programados

            if denominador_porcentaje > 0:
                porcentaje = round((numerador_porcentaje / denominador_porcentaje) * 100)
            else:
                porcentaje = 100

            fila_data += [d_atiempo, d_tardanzas, total_justif_perm_columna, d_inasistencias, f"{porcentaje}%"]
            data_para_tabla.append(fila_data)
            
            g_pdf_atiempo += d_atiempo
            g_pdf_tardanzas += d_tardanzas
            g_pdf_inasistencias += d_inasistencias
            g_pdf_justif_y_perm += total_justif_perm_columna

        pie_pdf = [""] * len(encabezado_principal_pdf)
        pie_pdf[1] = "TOTALES"
        pie_pdf[-5] = str(g_pdf_atiempo)
        pie_pdf[-4] = str(g_pdf_tardanzas)
        pie_pdf[-3] = str(g_pdf_justif_y_perm)
        pie_pdf[-2] = str(g_pdf_inasistencias)
        data_para_tabla.append(pie_pdf)

         
        reportes_dir_pdf = current_app.config['REPORTS_PATH']
        os.makedirs(reportes_dir_pdf, exist_ok=True)
        path_pdf = os.path.join(reportes_dir_pdf, f"reporte_asistencias_mensual_{mes}.pdf")

        doc = SimpleDocTemplate(path_pdf, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
        
         
        styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
        styles.add(ParagraphStyle(name='Left', alignment=TA_LEFT))
         
        styles['h2'].alignment = TA_CENTER
        
        titulo_pdf = Paragraph("PARTE MENSUAL DE ASISTENCIA DEL PERSONAL DOCENTE - PRIMARIA", styles['h2'])
        
        mes_dt = datetime.strptime(mes, '%Y-%m')
        meses_es = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        mes_nombre = meses_es[mes_dt.month - 1].upper()
        año_str = mes_dt.year
        info_izquierda_str = 'I.E. N° 82008 "SANTA BEATRIZ DE SILVA"- CAJAMARCA'
        info_derecha_str = f"MES: {mes_nombre}<br/>AÑO: {año_str}<br/>DIRECTOR: {director_nombre_completo}"
        p_info_izquierda = Paragraph(info_izquierda_str, styles['Normal'])
        p_info_derecha = Paragraph(info_derecha_str, styles['Normal'])
        tabla_encabezado = Table([[p_info_izquierda, p_info_derecha]], colWidths=['65%', '35%'])
        tabla_encabezado.setStyle(TableStyle([('BOX', (0,0), (-1,-1), 1, colors.black),('VALIGN', (0,0), (-1,-1), 'TOP'),('PADDING', (0,0), (-1,-1), 6)]))
        
        ancho_pagina, alto_pagina = landscape(A4)
        ancho_util = ancho_pagina - doc.leftMargin - doc.rightMargin
        ancho_nro = 22
        ancho_nombre = 150
        ancho_cargo = 80
        ancho_totales_grupo = 35
        ancho_fijo = ancho_nro + ancho_nombre + ancho_cargo + (ancho_totales_grupo * 5)
        ancho_restante_para_dias = ancho_util - ancho_fijo
        ancho_por_dia = ancho_restante_para_dias / len(fechas_del_mes) if fechas_del_mes else 15
        col_widths = [ancho_nro, ancho_nombre, ancho_cargo] + [ancho_por_dia] * len(fechas_del_mes) + [ancho_totales_grupo] * 5
        
        tabla_principal = Table(data_para_tabla, repeatRows=1, colWidths=col_widths)
        tabla_principal.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(" 
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 1), (2, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(" 
            ('FONTSIZE', (0, 0), (-1, -1), 6.5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements = [titulo_pdf, Spacer(1, 12), tabla_encabezado, Spacer(1, 12), tabla_principal]
        doc.build(elements)
        
        flash(f"✅ Reporte PDF generado en: {path_pdf}", "success")
    except Exception as e:
        current_app.logger.error(f"Error al generar PDF: {e.__class__.__name__}: {e}")
        import traceback
        traceback.print_exc()
        flash(f"❌ Ocurrió un error al generar el archivo PDF: {e.__class__.__name__}: {e}", "danger")

    return redirect(url_for('reports.reporte_mensual'))

@reports_bp.route('/dashboard')
@requiere_rol(1,2)
def dashboard():
    db = get_db()
    c = db.cursor()
    
    hoy_str = datetime.now().strftime('%Y-%m-%d')
    mes_actual_str = datetime.now().strftime('%Y-%m')
    dia_semana_hoy = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado", "Domingo"][datetime.now().weekday()]

     
    total_docentes = c.execute("SELECT COUNT(*) FROM docentes").fetchone()[0]
    total_justif_aprob = c.execute("SELECT COUNT(*) FROM justificaciones WHERE estado='Aprobada'").fetchone()[0]
    total_justif_pend = c.execute("SELECT COUNT(*) FROM justificaciones WHERE estado='Pendiente'").fetchone()[0]
    
    permisos_activos_hoy = c.execute("""
        SELECT COUNT(*) FROM permisos_docentes WHERE ? BETWEEN fecha_inicio AND fecha_fin
    """, (hoy_str,)).fetchone()[0]

     
    total_advertencias_mes = c.execute("SELECT COUNT(*) FROM advertencias WHERE strftime('%Y-%m', fecha) = ?", (mes_actual_str,)).fetchone()[0]

     
    asistencias_mes = c.execute("SELECT COUNT(*) FROM asistencias WHERE estado='A tiempo' AND strftime('%Y-%m', fecha)=?", (mes_actual_str,)).fetchone()[0]
    
    tardanzas_mes = c.execute("""
        SELECT COUNT(a.id) FROM asistencias a
        LEFT JOIN justificaciones j ON a.id = j.asistencia_id AND j.estado = 'Aprobada'
        WHERE a.estado = 'Tarde' AND j.id IS NULL AND strftime('%Y-%m', a.fecha) = ?
    """, (mes_actual_str,)).fetchone()[0]
    
    inasistencias_mes = c.execute("""
        SELECT COUNT(a.id) FROM asistencias a
        LEFT JOIN justificaciones j ON a.id = j.asistencia_id AND j.estado = 'Aprobada'
        WHERE a.estado = 'Inasistencia' AND j.id IS NULL AND strftime('%Y-%m', a.fecha) = ?
    """, (mes_actual_str,)).fetchone()[0]


     
    docentes_programados_hoy = c.execute("""
        SELECT COUNT(DISTINCT d.dni) FROM docentes d
        JOIN horarios h ON d.dni = h.dni
        WHERE h.dia_semana = ? AND d.dni NOT IN (
            SELECT p.dni FROM permisos_docentes p WHERE ? BETWEEN p.fecha_inicio AND p.fecha_fin
        )
    """, (dia_semana_hoy, hoy_str)).fetchone()[0]

    asistencias_reales_hoy = c.execute("""
        SELECT COUNT(*) FROM asistencias WHERE DATE(fecha) = ? AND estado IN ('A tiempo', 'Tarde')
    """, (hoy_str,)).fetchone()[0]
    
    porcentaje_asistieron_hoy = round((asistencias_reales_hoy / docentes_programados_hoy) * 100, 1) if docentes_programados_hoy > 0 else 0

    estado_docentes_hoy = c.execute("""
        WITH DocentesProgramados AS (
            SELECT d.dni, d.nombres || ' ' || d.apellido_paterno AS nombre
            FROM docentes d JOIN horarios h ON d.dni = h.dni
            WHERE h.dia_semana = ?
        )
        SELECT
            dp.nombre,
            COALESCE(a.estado, p.estado_permiso, 'Pendiente') AS estado,
            a.hora_registro,
            j.estado AS estado_justificacion
        FROM DocentesProgramados dp
        LEFT JOIN asistencias a ON dp.dni = a.dni AND DATE(a.fecha) = ?
        LEFT JOIN justificaciones j ON a.id = j.asistencia_id
        LEFT JOIN (
            SELECT dni, 'Permiso' as estado_permiso FROM permisos_docentes WHERE ? BETWEEN fecha_inicio AND fecha_fin
        ) p ON dp.dni = p.dni
        ORDER BY
            CASE WHEN COALESCE(a.estado, p.estado_permiso, 'Pendiente') = 'A tiempo' THEN 1
                 WHEN COALESCE(a.estado, p.estado_permiso, 'Pendiente') = 'Tarde' THEN 2
                 WHEN COALESCE(a.estado, p.estado_permiso, 'Pendiente') = 'Permiso' THEN 3
                 WHEN COALESCE(a.estado, p.estado_permiso, 'Pendiente') = 'Pendiente' THEN 4
                 ELSE 5 END,
            a.hora_registro
    """, (dia_semana_hoy, hoy_str, hoy_str)).fetchall()


     
    docente_destacado = c.execute("""
        SELECT d.nombres || ' ' || d.apellido_paterno AS nombre, COUNT(a.id) AS cantidad
        FROM asistencias a JOIN docentes d ON a.dni = d.dni
        WHERE a.estado = 'A tiempo' AND strftime('%Y-%m', a.fecha) = ?
        GROUP BY a.dni
        HAVING cantidad > 0
        ORDER BY cantidad DESC LIMIT 1
    """, (mes_actual_str,)).fetchone()

    top_puntuales = c.execute("""
        WITH Stats AS (
            SELECT
                a.dni,
                SUM(CASE WHEN a.estado = 'A tiempo' THEN 1 ELSE 0 END) AS a_tiempo,
                SUM(CASE WHEN a.estado = 'Tarde' AND j.id IS NULL THEN 1 ELSE 0 END) AS tardanzas,
                SUM(CASE WHEN a.estado = 'Inasistencia' AND j.id IS NULL THEN 1 ELSE 0 END) AS inasistencias
            FROM asistencias a
            LEFT JOIN justificaciones j ON a.id = j.asistencia_id AND j.estado = 'Aprobada'
            WHERE strftime('%Y-%m', a.fecha) = ?
            GROUP BY a.dni
        )
        SELECT
            d.nombres || ' ' || d.apellido_paterno AS nombre,
            ROUND( (s.a_tiempo * 100.0) / (s.a_tiempo + s.tardanzas + s.inasistencias) , 1) AS porcentaje
        FROM Stats s JOIN docentes d ON s.dni = d.dni
        WHERE (s.a_tiempo + s.tardanzas + s.inasistencias) > 0 AND s.a_tiempo > 0
        ORDER BY porcentaje DESC, s.a_tiempo DESC
        LIMIT 5
    """, (mes_actual_str,)).fetchall()

    tardanzas_docentes = c.execute("""
        SELECT d.nombres||' '||d.apellido_paterno AS nombre, COUNT(*) AS cantidad
        FROM asistencias a
        JOIN docentes d ON d.dni=a.dni
        LEFT JOIN justificaciones j ON j.asistencia_id=a.id AND j.estado='Aprobada'
        WHERE a.estado='Tarde' AND j.id IS NULL AND strftime('%Y-%m', a.fecha) = ?
        GROUP BY a.dni
        ORDER BY cantidad DESC LIMIT 5
    """, (mes_actual_str,)).fetchall()
    
     
    top_inasistencias = c.execute("""
        SELECT d.nombres||' '||d.apellido_paterno AS nombre, COUNT(*) AS cantidad
        FROM asistencias a
        JOIN docentes d ON d.dni=a.dni
        LEFT JOIN justificaciones j ON j.asistencia_id=a.id AND j.estado='Aprobada'
        WHERE a.estado='Inasistencia' AND j.id IS NULL AND strftime('%Y-%m', a.fecha) = ?
        GROUP BY a.dni
        ORDER BY cantidad DESC LIMIT 5
    """, (mes_actual_str,)).fetchall()

     
    top_advertencias = c.execute("""
        SELECT d.nombres || ' ' || d.apellido_paterno AS nombre, COUNT(a.id) AS cantidad
        FROM advertencias a
        JOIN docentes d ON a.dni = d.dni
        WHERE strftime('%Y-%m', a.fecha) = ?
        GROUP BY a.dni
        ORDER BY cantidad DESC
        LIMIT 5
    """, (mes_actual_str,)).fetchall()

     
    justificadas_mes = c.execute("""
        SELECT COUNT(j.id) FROM justificaciones j
        JOIN asistencias a ON j.asistencia_id = a.id
        WHERE j.estado = 'Aprobada' AND strftime('%Y-%m', a.fecha) = ?
    """, (mes_actual_str,)).fetchone()[0]
    
    permisos_mes = c.execute("SELECT COUNT(DISTINCT d.dni) FROM permisos_docentes p JOIN docentes d ON p.dni = d.dni WHERE strftime('%Y-%m', p.fecha_inicio) = ? OR strftime('%Y-%m', p.fecha_fin) = ?", (mes_actual_str, mes_actual_str)).fetchone()[0]

    puntualidad_general_data = {
        'a_tiempo': asistencias_mes,
        'tardanzas': tardanzas_mes,
        'inasistencias': inasistencias_mes,
        'justificadas_permisos': justificadas_mes + permisos_mes
    }

    fechas_7dias = [(datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(6, -1, -1)]
    asistencias_7dias_data = []
    for fecha in fechas_7dias:
        dia_semana_dt = datetime.strptime(fecha, '%Y-%m-%d')
        if dia_semana_dt.weekday() < 5: 
            counts = c.execute("""
                SELECT
                    SUM(CASE WHEN estado = 'A tiempo' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN estado = 'Tarde' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN estado = 'Inasistencia' THEN 1 ELSE 0 END)
                FROM asistencias WHERE DATE(fecha) = ?
            """, (fecha,)).fetchone()
            
            permisos_dia = c.execute("SELECT COUNT(*) FROM permisos_docentes WHERE ? BETWEEN fecha_inicio AND fecha_fin", (fecha,)).fetchone()[0]

            asistencias_7dias_data.append({
                'fecha': dia_semana_dt.strftime('%d-%b'),
                'a_tiempo': counts[0] or 0,
                'tardanzas': counts[1] or 0,
                'inasistencias': counts[2] or 0,
                'permisos': permisos_dia
            })

     
    return render_template('dashboard.html',
         
        total_docentes=total_docentes,
        asistencias_mes=asistencias_mes,
        tardanzas_mes=tardanzas_mes,
        inasistencias_mes=inasistencias_mes,
        total_justif_aprob=total_justif_aprob,
        total_justif_pend=total_justif_pend,
        permisos_activos_hoy=permisos_activos_hoy,
        total_advertencias_mes=total_advertencias_mes,  
        
         
        porcentaje_asistieron_hoy=porcentaje_asistieron_hoy,
        docentes_programados_hoy=docentes_programados_hoy,
        estado_docentes_hoy=estado_docentes_hoy,
        
         
        docente_destacado=docente_destacado,
        top_puntuales=top_puntuales,
        tardanzas_docentes=tardanzas_docentes,
        top_inasistencias=top_inasistencias,  
        top_advertencias=top_advertencias,    

         
        puntualidad_general_data=puntualidad_general_data,
        asistencias_7dias_data=asistencias_7dias_data
    )

